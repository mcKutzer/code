"""Download monthly price history for the MPT asset universe, compute
annualised return / volatility per asset (using each asset's own full
history) and pairwise correlations (using overlapping months only), and
write everything to data/inputs.xlsx for review/override.

Run directly to (re)build data/inputs.xlsx and print a summary.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT: Path = Path(__file__).resolve().parent.parent
ETF_TABLE_PATH: Path = ROOT / "ETFs.xlsx"
OUT_PATH: Path = ROOT / "data" / "inputs.xlsx"

PERIOD: str = "max"
INTERVAL: str = "1mo"


# ---------- Universe -----------------------------------------------------------


def load_universe() -> pd.DataFrame:
    """Read ETFs.xlsx — already curated, tickers are yfinance-ready."""
    df = pd.read_excel(ETF_TABLE_PATH)
    df = df.rename(columns=lambda c: str(c).strip())
    required = {"Asset class", "ETF", "Ticker"}
    missing = required - set(df.columns)
    if missing:
        print(f"[fetch_prices] ERROR: ETFs.xlsx missing columns: {missing}")
        sys.exit(1)
    return df.reset_index(drop=True)


# ---------- Download -----------------------------------------------------------


def download_prices(yf_tickers: list[str]) -> pd.DataFrame:
    """Monthly close prices, one column per yf ticker. NaNs preserved for
    months before an ETF existed."""
    raw = yf.download(
        tickers=yf_tickers,
        period=PERIOD,
        interval=INTERVAL,
        auto_adjust=True,
        progress=False,
        group_by="ticker",
        threads=True,
    )
    if raw is None or raw.empty:
        print("[fetch_prices] ERROR: yfinance returned no data.")
        sys.exit(1)

    if isinstance(raw.columns, pd.MultiIndex):
        frames = {}
        for t in yf_tickers:
            if t in raw.columns.get_level_values(0) and "Close" in raw[t].columns:
                frames[t] = raw[t]["Close"]
        prices = pd.DataFrame(frames)
    else:
        prices = raw[["Close"]].rename(columns={"Close": yf_tickers[0]})

    return prices[[c for c in yf_tickers if c in prices.columns]]


# ---------- Stats --------------------------------------------------------------


def compute_returns(prices: pd.DataFrame) -> pd.DataFrame:
    """Monthly log returns. NaNs propagate where price data is absent."""
    return np.log(prices / prices.shift(1))


def per_asset_stats(log_returns: pd.DataFrame) -> pd.DataFrame:
    """For each column, drop NaNs and compute annualised mean / stdev plus
    inception date and obs count from that asset's own history."""
    rows = []
    for col in log_returns.columns:
        s = log_returns[col].dropna()
        if s.empty:
            rows.append({"Ticker": col, "FirstDate": None, "LastDate": None,
                         "NObs": 0, "AnnReturn_calc": np.nan, "AnnVol_calc": np.nan})
            continue
        rows.append({
            "Ticker": col,
            "FirstDate": s.index.min().date(),
            "LastDate": s.index.max().date(),
            "NObs": int(s.size),
            "AnnReturn_calc": float(s.mean() * 12.0),
            "AnnVol_calc": float(s.std(ddof=1) * np.sqrt(12.0)),
        })
    return pd.DataFrame(rows)


def pairwise_correlation(log_returns: pd.DataFrame) -> pd.DataFrame:
    """pandas DataFrame.corr() uses pairwise complete observations by
    default — exactly what we want when histories don't all start together."""
    return log_returns.corr(method="pearson", min_periods=12)


# ---------- Excel output -------------------------------------------------------


def _autofit(ws, df: pd.DataFrame, extra: int = 2) -> None:
    for i, col in enumerate(df.columns, start=1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col].astype(str)))
        ws.column_dimensions[get_column_letter(i)].width = min(width + extra, 40)


def write_excel(
    universe: pd.DataFrame,
    stats: pd.DataFrame,
    corr: pd.DataFrame,
) -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # ---- Sheet 1: Returns ----
    merged = universe.merge(stats, on="Ticker", how="left")
    cols = {
        "Asset class": merged["Asset class"],
        "ETF": merged["ETF"],
        "Ticker": merged["Ticker"],
        "First date": merged["FirstDate"],
        "Last date": merged["LastDate"],
        "Months": merged["NObs"],
        "Ann Return (calc)": merged["AnnReturn_calc"],
        "Ann Vol (calc)": merged["AnnVol_calc"],
        "Ann Return (override)": [None] * len(merged),
        "Ann Vol (override)": [None] * len(merged),
    }
    if "Baseline" in merged.columns:
        cols["Baseline"] = merged["Baseline"]
    returns_df = pd.DataFrame(cols)

    # ---- Sheet 2: Correlation matrix ordered by asset class ----
    ordered_tickers = universe["Ticker"].tolist()
    ordered_names = universe["Asset class"].tolist()
    corr_ord = corr.reindex(index=ordered_tickers, columns=ordered_tickers)
    corr_ord.index = ordered_names
    corr_ord.columns = ordered_names

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        returns_df.to_excel(writer, sheet_name="Returns", index=False)
        corr_ord.to_excel(writer, sheet_name="Correlation", index=True)

        # ---- Formatting: Returns sheet ----
        ws_r = writer.sheets["Returns"]
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF1F2A44")
        override_fill = PatternFill("solid", fgColor="FFFFF7C2")
        for col_idx, _ in enumerate(returns_df.columns, start=1):
            cell = ws_r.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Percentage format on calc + override return/vol columns
        col_idx_of = {name: i + 1 for i, name in enumerate(returns_df.columns)}
        pct_cols = [
            col_idx_of["Ann Return (calc)"],
            col_idx_of["Ann Vol (calc)"],
            col_idx_of["Ann Return (override)"],
            col_idx_of["Ann Vol (override)"],
        ]
        for col in pct_cols:
            for row in range(2, len(returns_df) + 2):
                ws_r.cell(row=row, column=col).number_format = "0.00%"
        # Highlight override columns so user knows where to type
        for col in (col_idx_of["Ann Return (override)"], col_idx_of["Ann Vol (override)"]):
            for row in range(2, len(returns_df) + 2):
                ws_r.cell(row=row, column=col).fill = override_fill
        if "Baseline" in col_idx_of:
            for row in range(2, len(returns_df) + 2):
                ws_r.cell(row=row, column=col_idx_of["Baseline"]).number_format = "0.00%"
        _autofit(ws_r, returns_df)
        ws_r.freeze_panes = "B2"

        # ---- Formatting: Correlation sheet (heatmap) ----
        ws_c = writer.sheets["Correlation"]
        n = len(ordered_names)
        # Header row + first column styling
        for col_idx in range(1, n + 2):
            ws_c.cell(row=1, column=col_idx).font = header_font
            ws_c.cell(row=1, column=col_idx).fill = header_fill
            ws_c.cell(row=1, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for row_idx in range(2, n + 2):
            ws_c.cell(row=row_idx, column=1).font = Font(bold=True)
            ws_c.cell(row=row_idx, column=1).fill = header_fill
            ws_c.cell(row=row_idx, column=1).font = Font(bold=True, color="FFFFFFFF")
        # Number format on the matrix body
        for row in range(2, n + 2):
            for col in range(2, n + 2):
                ws_c.cell(row=row, column=col).number_format = "0.00"
                ws_c.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        # Heatmap: red (-1) -> white (0) -> green (+1)
        body_range = f"B2:{get_column_letter(n + 1)}{n + 1}"
        ws_c.conditional_formatting.add(
            body_range,
            ColorScaleRule(
                start_type="num", start_value=-1, start_color="FFE06B6B",
                mid_type="num", mid_value=0, mid_color="FFFFFFFF",
                end_type="num", end_value=1, end_color="FF2E8B57",
            ),
        )
        # Width: first column wider for asset names, others uniform
        ws_c.column_dimensions["A"].width = 28
        for col_idx in range(2, n + 2):
            ws_c.column_dimensions[get_column_letter(col_idx)].width = 14
        ws_c.row_dimensions[1].height = 60
        ws_c.freeze_panes = "B2"


# ---------- Orchestration ------------------------------------------------------


def _print_summary(returns_df: pd.DataFrame, corr: pd.DataFrame) -> None:
    print("\nPer-asset stats (each on its own full history):\n")
    show = returns_df[["Asset class", "Ticker", "FirstDate",
                       "NObs", "AnnReturn_calc", "AnnVol_calc"]].copy()
    show["AnnReturn_calc"] = show["AnnReturn_calc"].apply(lambda x: f"{x:.2%}" if pd.notna(x) else "n/a")
    show["AnnVol_calc"] = show["AnnVol_calc"].apply(lambda x: f"{x:.2%}" if pd.notna(x) else "n/a")
    print(show.to_string(index=False))
    print(f"\nCorrelation matrix: {corr.shape[0]} x {corr.shape[1]} "
          f"(pairwise complete observations)")
    print(f"\nSaved -> {OUT_PATH}")


def main() -> None:
    universe = load_universe()
    print(f"Universe: {len(universe)} ETFs\n")
    print(universe[["Asset class", "Ticker", "ETF"]].to_string(index=False))

    prices = download_prices(universe["Ticker"].tolist())
    log_returns = compute_returns(prices)
    stats = per_asset_stats(log_returns)
    corr = pairwise_correlation(log_returns)

    merged = universe.merge(stats, on="Ticker", how="left")
    write_excel(universe, stats, corr)
    _print_summary(merged, corr)


if __name__ == "__main__":
    main()
