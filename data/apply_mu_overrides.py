"""Populate the 'Ann Return (override)' column in inputs.xlsx with forward-
looking capital market assumptions (Vanguard/BlackRock CMA 2024).

Keyed by ticker. Only the 'Ann Return (override)' column is touched; every
other cell (calc values, vol overrides, baseline, correlation/covariance
sheets) is left alone.

Re-run any time to reset overrides to the CMA values.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT: Path = Path(__file__).resolve().parent.parent
INPUTS_PATH: Path = ROOT / "data" / "inputs.xlsx"

# Forward-looking expected returns (annualised, EUR). Equities + gold + REITs +
# commodities from Vanguard/BlackRock CMA 2024; EUR bond rows aligned to the
# current EUR sovereign/IG yield curve; cash anchored to the ECB deposit rate.
CMA_MU: dict[str, float] = {
    "VUAA.DE":   0.105,   # S&P 500            (USD equity, CMA)
    "EXSA.DE":   0.085,   # Europe equity      (CMA)
    "IS3N.DE":   0.100,   # Emerging markets   (CMA)
    "CSBGE3.MI": 0.027,   # Short gov EUR      (1-3yr EUR sovereign yield)
    "SXRP.DE":   0.030,   # Med gov EUR        (3-7yr EUR sovereign yield)
    "IEAA.L":    0.036,   # Corp bonds IG EUR  (base + IG spread)
    "IGBY.AS":   0.040,   # Long gov EUR       (15-30yr EUR sovereign yield)
    "XAD5.MI":   0.065,   # Gold               (CMA)
    "IQQ6.DE":   0.075,   # Global REITs       (CMA)
    "XEON.DE":   0.020,   # Cash equivalent    (ECB deposit rate)
    "IHYG.L":    0.0554,  # High Yield Bonds EUR (HY spread + base, user input)
}

SHEET_NAME: str = "Returns"
TICKER_HEADER: str = "Ticker"
OVERRIDE_HEADER: str = "Ann Return (override)"


def main() -> None:
    if not INPUTS_PATH.exists():
        print(f"[apply_mu_overrides] ERROR: {INPUTS_PATH} not found.")
        sys.exit(1)

    wb = load_workbook(INPUTS_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[apply_mu_overrides] ERROR: sheet '{SHEET_NAME}' missing.")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    # Resolve column indices from the header row
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    if TICKER_HEADER not in header or OVERRIDE_HEADER not in header:
        print(f"[apply_mu_overrides] ERROR: missing columns. Found: {list(header)}")
        sys.exit(1)
    tcol = header[TICKER_HEADER]
    ocol = header[OVERRIDE_HEADER]

    applied: list[str] = []
    unknown: list[str] = []
    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=tcol).value
        if not ticker:
            continue
        if ticker in CMA_MU:
            ws.cell(row=row, column=ocol, value=CMA_MU[ticker])
            ws.cell(row=row, column=ocol).number_format = "0.00%"
            applied.append(ticker)
        else:
            unknown.append(str(ticker))

    missing = [t for t in CMA_MU if t not in applied]

    try:
        wb.save(INPUTS_PATH)
    except PermissionError:
        print(f"[apply_mu_overrides] ERROR: {INPUTS_PATH} is open in Excel. Close it and re-run.")
        sys.exit(1)

    print(f"Applied CMA mu overrides to {len(applied)} rows:")
    for t in applied:
        print(f"  {t:<10s} -> {CMA_MU[t]:.2%}")
    if missing:
        print(f"\nNot found in sheet (CMA entries with no matching ticker): {missing}")
    if unknown:
        print(f"Rows in sheet with no CMA value (left unchanged): {unknown}")
    print(f"\nSaved -> {INPUTS_PATH}")


if __name__ == "__main__":
    main()
