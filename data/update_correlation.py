"""Recompute the correlation matrix on EUR-denominated log returns.

For each asset class we pick a long-history source ticker (preferring liquid
US listings for equities / gold), auto-detect its trading currency from
yfinance, pull the EUR/<ccy> FX series if needed, convert the price series
to EUR, and compute pairwise correlations using complete observations only.

Updates ONLY the 'Correlation' sheet of data/inputs.xlsx — the Returns
sheet (and any user overrides there) is left untouched.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT: Path = Path(__file__).resolve().parent.parent
ETF_TABLE_PATH: Path = ROOT / "ETFs.xlsx"
INPUTS_PATH: Path = ROOT / "data" / "inputs.xlsx"

# asset class -> (correlation source ticker, note for log)
# Default is the ETFs.xlsx ticker; entries below override it where a
# longer-history liquid proxy with identical underlying exposure exists.
SOURCE_OVERRIDES: dict[str, str] = {
    "S&P 500": "SPY",                       # USD, since 1993
    "Emerging markets": "EEM",              # USD, since 2003
    "Long gov bonds": "IDTL.L",             # GBP, same iShares fund as IGBY.AS, longer history
    "Gold": "GLD",                          # USD, since 2004
    "High Yield Bonds EUR": "HYG",          # USD, since 2007; US HY ~0.85 corr with EUR HY
}

PERIOD: str = "max"
INTERVAL: str = "1mo"

FX_TICKER: dict[str, str] = {
    "USD": "EURUSD=X",   # USD per 1 EUR
    "GBP": "EURGBP=X",   # GBP per 1 EUR
}

EWMA_HALFLIFE_MONTHS: int = 36   # ~ lambda = 0.94, RiskMetrics standard
EWMA_MIN_PERIODS: int = 12
SAMPLE_MIN_PERIODS: int = 12


# ---------- Universe -----------------------------------------------------------


def build_source_map() -> pd.DataFrame:
    """Return a DataFrame with columns [Asset class, OriginalTicker, SourceTicker]."""
    df = pd.read_excel(ETF_TABLE_PATH).rename(columns=lambda c: str(c).strip())
    df["SourceTicker"] = df.apply(
        lambda r: SOURCE_OVERRIDES.get(r["Asset class"], r["Ticker"]),
        axis=1,
    )
    return df[["Asset class", "Ticker", "SourceTicker"]].rename(
        columns={"Ticker": "OriginalTicker"}
    )


# ---------- Download & FX -----------------------------------------------------


def _detect_currency(ticker: str) -> str:
    """Look up the trading currency via yfinance. Falls back to 'EUR' on
    failure (logged so the user can spot it)."""
    try:
        info = yf.Ticker(ticker).fast_info
        ccy = info.get("currency") if hasattr(info, "get") else info.currency
        if ccy:
            return str(ccy).upper()
    except Exception as e:
        print(f"  [warn] currency lookup failed for {ticker}: {e}")
    print(f"  [warn] no currency found for {ticker}; assuming EUR")
    return "EUR"


def download_monthly(ticker: str) -> pd.Series:
    h = yf.Ticker(ticker).history(period=PERIOD, interval=INTERVAL, auto_adjust=True)
    if h.empty:
        print(f"  [warn] no data for {ticker}")
        return pd.Series(dtype=float, name=ticker)
    s = h["Close"].copy()
    s.index = s.index.tz_localize(None).to_period("M").to_timestamp(how="end").normalize()
    s.name = ticker
    return s


def download_fx(currency: str) -> pd.Series | None:
    """EUR per 1 unit of `currency`? No — yfinance EURUSD=X quotes USD per 1 EUR,
    so to convert a USD-priced asset to EUR you divide by EURUSD=X."""
    if currency == "EUR":
        return None
    fx_ticker = FX_TICKER.get(currency)
    if not fx_ticker:
        print(f"  [warn] no FX mapping for {currency}; series left in {currency}")
        return None
    fx = download_monthly(fx_ticker)
    if fx.empty:
        print(f"  [warn] FX series {fx_ticker} empty")
        return None
    return fx


def to_eur(price: pd.Series, currency: str) -> pd.Series:
    if currency == "EUR":
        return price
    fx = download_fx(currency)
    if fx is None:
        return price
    aligned_fx = fx.reindex(price.index, method="nearest", tolerance=pd.Timedelta(days=20))
    converted = price / aligned_fx
    converted.name = price.name
    return converted.dropna()


# ---------- Correlation -------------------------------------------------------


def _corr_from_cov(cov: pd.DataFrame) -> pd.DataFrame:
    sigma = np.sqrt(np.diag(cov.values))
    inv = np.where(sigma > 0, 1.0 / sigma, 0.0)
    corr = cov.values * np.outer(inv, inv)
    np.fill_diagonal(corr, 1.0)
    return pd.DataFrame(corr, index=cov.index, columns=cov.columns)


def build_matrices(
    srcmap: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Returns (corr_sample, corr_ewma, cov_sample, cov_ewma, info).
    All matrices are annualised and indexed/columned by Asset class name."""
    eur_prices: dict[str, pd.Series] = {}
    info_rows = []

    for _, row in srcmap.iterrows():
        ac = row["Asset class"]
        src = row["SourceTicker"]
        orig = row["OriginalTicker"]
        print(f"- {ac:<20s} source={src}" + (f"  (was {orig})" if src != orig else ""))
        ccy = _detect_currency(src)
        price = download_monthly(src)
        if price.empty:
            print(f"  [skip] {src} returned empty")
            continue
        eur = to_eur(price, ccy)
        eur_prices[ac] = eur
        info_rows.append({
            "Asset class": ac,
            "SourceTicker": src,
            "Currency": ccy,
            "FirstDate": eur.index.min().date(),
            "LastDate": eur.index.max().date(),
            "Months": len(eur),
        })

    if not eur_prices:
        print("[update_correlation] ERROR: no series collected.")
        sys.exit(1)

    prices_df = pd.DataFrame(eur_prices).sort_index()
    log_returns = np.log(prices_df / prices_df.shift(1))
    order = [ac for ac in srcmap["Asset class"] if ac in log_returns.columns]
    log_returns = log_returns[order]

    # Sample covariance (annualised, pairwise complete obs)
    cov_sample = log_returns.cov(min_periods=SAMPLE_MIN_PERIODS) * 12.0

    # EWMA covariance: ewm(...).cov() returns a (date, asset) x asset multi-index.
    # We want the last available cov matrix (most-recent estimate).
    ewm_cov = log_returns.ewm(
        halflife=EWMA_HALFLIFE_MONTHS, min_periods=EWMA_MIN_PERIODS
    ).cov()
    last_date = log_returns.dropna(how="all").index.max()
    cov_ewma = ewm_cov.xs(last_date, level=0).reindex(index=order, columns=order) * 12.0

    corr_sample = _corr_from_cov(cov_sample)
    corr_ewma = _corr_from_cov(cov_ewma)

    return corr_sample, corr_ewma, cov_sample, cov_ewma, pd.DataFrame(info_rows)


# ---------- Excel write (Correlation sheet only) ------------------------------


def _format_correlation_sheet(ws, n: int) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F2A44")
    for col_idx in range(1, n + 2):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(2, n + 2):
        c = ws.cell(row=row_idx, column=1)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = header_fill
    for row in range(2, n + 2):
        for col in range(2, n + 2):
            ws.cell(row=row, column=col).number_format = "0.00"
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    body_range = f"B2:{get_column_letter(n + 1)}{n + 1}"
    ws.conditional_formatting.add(
        body_range,
        ColorScaleRule(
            start_type="num", start_value=-1, start_color="FFE06B6B",
            mid_type="num", mid_value=0, mid_color="FFFFFFFF",
            end_type="num", end_value=1, end_color="FF2E8B57",
        ),
    )
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, n + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.row_dimensions[1].height = 60
    ws.freeze_panes = "B2"


def _write_matrix_sheet(
    wb,
    sheet_name: str,
    df: pd.DataFrame,
    heatmap: bool,
    number_format: str,
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value="Asset class")
    for j, name in enumerate(df.columns, start=2):
        ws.cell(row=1, column=j, value=name)
    for i, name in enumerate(df.index, start=2):
        ws.cell(row=i, column=1, value=name)
        for j, _ in enumerate(df.columns, start=2):
            v = df.iloc[i - 2, j - 2]
            ws.cell(row=i, column=j, value=None if pd.isna(v) else float(v))
    n = len(df.columns)

    if heatmap:
        _format_correlation_sheet(ws, n=n)
    else:
        # Covariance sheet: header strip + number format + freeze panes
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF1F2A44")
        for col_idx in range(1, n + 2):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_idx in range(2, n + 2):
            c = ws.cell(row=row_idx, column=1)
            c.font = Font(bold=True, color="FFFFFFFF")
            c.fill = header_fill
        for row in range(2, n + 2):
            for col in range(2, n + 2):
                ws.cell(row=row, column=col).number_format = number_format
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        for col_idx in range(2, n + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        ws.row_dimensions[1].height = 60
        ws.freeze_panes = "B2"


def write_matrices_to_workbook(
    corr_sample: pd.DataFrame,
    corr_ewma: pd.DataFrame,
    cov_sample: pd.DataFrame,
    cov_ewma: pd.DataFrame,
    info: pd.DataFrame,
) -> None:
    if not INPUTS_PATH.exists():
        print(f"[update_correlation] ERROR: {INPUTS_PATH} not found. "
              "Run data/fetch_prices.py first.")
        sys.exit(1)

    wb = load_workbook(INPUTS_PATH)

    # Correlation (sample) — the heatmap the user has been seeing
    _write_matrix_sheet(wb, "Correlation", corr_sample, heatmap=True, number_format="0.00")
    # EWMA correlation — second heatmap for comparison
    _write_matrix_sheet(wb, "Correlation_EWMA", corr_ewma, heatmap=True, number_format="0.00")
    # Annualised covariance matrices — used by the optimizer via cov_method
    _write_matrix_sheet(wb, "Covariance_Sample", cov_sample, heatmap=False, number_format="0.0000")
    _write_matrix_sheet(wb, "Covariance_EWMA", cov_ewma, heatmap=False, number_format="0.0000")

    # CorrelationSources (rebuild)
    if "CorrelationSources" in wb.sheetnames:
        del wb["CorrelationSources"]
    ws2 = wb.create_sheet("CorrelationSources")
    for j, col in enumerate(info.columns, start=1):
        ws2.cell(row=1, column=j, value=col)
        ws2.cell(row=1, column=j).font = Font(bold=True, color="FFFFFFFF")
        ws2.cell(row=1, column=j).fill = PatternFill("solid", fgColor="FF1F2A44")
    for i, row in enumerate(info.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=v)
    for col_idx, col in enumerate(info.columns, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(16, len(col) + 2)
    ws2.freeze_panes = "A2"

    wb.save(INPUTS_PATH)


# ---------- Main --------------------------------------------------------------


def main() -> None:
    srcmap = build_source_map()
    print(f"Universe: {len(srcmap)} asset classes\n")

    corr_sample, corr_ewma, cov_sample, cov_ewma, info = build_matrices(srcmap)

    print(f"\nCorrelation/covariance matrices: {corr_sample.shape[0]} x {corr_sample.shape[1]} "
          f"(EUR returns, annualised)\n")
    print(f"EWMA half-life: {EWMA_HALFLIFE_MONTHS} months\n")
    print("Source / currency / window:")
    show = info.copy()
    show["Months"] = show["Months"].astype(int)
    print(show.to_string(index=False))

    write_matrices_to_workbook(corr_sample, corr_ewma, cov_sample, cov_ewma, info)
    print(f"\nUpdated -> {INPUTS_PATH}")
    print("  sheets: Correlation, Correlation_EWMA, Covariance_Sample, "
          "Covariance_EWMA, CorrelationSources")


if __name__ == "__main__":
    main()
